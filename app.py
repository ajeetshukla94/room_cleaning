import pandas as pd
from flask import Flask, render_template, flash, url_for, request, make_response, jsonify, session,send_from_directory
from werkzeug.utils import secure_filename
import os, time
import io
import base64
import json
import datetime
import xlsxwriter
import os, sys, glob
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Border, Side
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment
from openpyxl.styles import Border, Side
from flask import send_file
import smtplib,ssl
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email.utils import formatdate
from email import encoders
from fnmatch import fnmatch
import time
from openpyxl.styles import PatternFill
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from sheet_generation import Sheet_Generation

app = Flask(__name__)
app.secret_key = 'file_upload_key'

MYDIR = os.path.dirname(__file__)
app.config['UPLOAD_FOLDER_REPORT'] = "static/Report/"
equipmet_list = ['Dispensing Booth',
'Dispensing Scoop ( Small)','Dispensing Scoop (Large)',
'Spatula','Spatula' ,'Manufacturing Scoop',
'Blender Bin (350 liter)' ,'Blender Bin (25 liter)' ,
'Vibro Sifter II','sampling rod','sampling rod' 
]

MYDIR = os.path.dirname(__file__)
app.config['UPLOAD_FOLDER_INPUTDATA'] = "static/inputData/"

sent_mail = False
server    = 'smtp.gmail.com'
port      =  587
username  =  "aajeetshk@gmail.com"
password  =  "ilbumnmnsnqletdk"
send_from = "aajeetshk@gmail.com"
send_to   = "ashish@pinpointengineers.co.in"

def send_mail(subject,text,files,file_name,isTls=True):
        msg = MIMEMultipart()
        msg['From'] = send_from
        msg['To'] = send_to
        msg['Date'] = formatdate(localtime = True)
        msg['Subject'] = subject
        msg.attach(MIMEText(text))

        part = MIMEBase('application', "octet-stream")
        part.set_payload(open(files, "rb").read())
        encoders.encode_base64(part)
        part.add_header('Content-Disposition', 'attachment; filename={}.xlsx'.format(file_name))
        msg.attach(part)
        smtp = smtplib.SMTP(server, port)
        if isTls:
            smtp.starttls()
        smtp.login(username,password)
        smtp.sendmail(send_from, send_to, msg.as_string())
        smtp.quit()
@app.route("/")
def default():
    return make_response(render_template('login_page/login.html'),200)    
    
@app.route("/login", methods=["GET", "POST"])
def login():
    product_frame  = pd.read_excel(os.path.join(app.config['UPLOAD_FOLDER_INPUTDATA'],"product_Details.xlsx"))
    product_list   = product_frame.Product_Name.unique().tolist()
    if request.method == 'POST':
      form_data = request.form
      l_id = form_data['login']
      pwd = form_data['password']
      if(l_id.lower()=='admin'.lower() and pwd == 'admin'):
          print('inside if')
          session['username'] = l_id
          flash('Login Successful')
          return make_response(render_template('cleaning_room.html',equipmet_list  = equipmet_list,product_list  = product_list,
                                msg = True, err = False, warn = False),200)
      else:
          print('inside else')
          flash('Invalid Credentials')
          return make_response(render_template("login_page/login.html", msg = False, err = True, warn = False),403)
    else:
        print('get request')
        
@app.route('/logout')
def logout():
    session.pop('username', None)
    session.pop('selected_file', None)
    global Selected_files
    Selected_file = None
    flash('Logout Successful')
    return make_response(render_template("login_page/login.html",msg = True, err = False, warn = False, message='Logout Successful'),200)

@app.route("/cleaning_room")
def cleaning_room():    
    product_frame  = pd.read_excel(os.path.join(app.config['UPLOAD_FOLDER_INPUTDATA'],"product_Details.xlsx"))
    product_list   = product_frame.Product_Name.unique().tolist()
    return make_response(render_template('cleaning_room.html',equipmet_list  = equipmet_list,product_list  = product_list),200) 
    
    
@app.route("/UpdateProductList")
def UpdateProductList():
    product_frame  = pd.read_excel(os.path.join(app.config['UPLOAD_FOLDER_INPUTDATA'],"product_Details.xlsx"))
    product_list   = product_frame.to_dict('records')
    return make_response(render_template('UpdateProductList.html',product_list  = product_list),200) 
    
@app.route("/submit_UpdateProductList")    
def submit_UpdateProductList():   
    data            = request.args.get('params_data')
    data            = json.loads(data)  
    observation     = data['observation']
    temp_df         = pd.DataFrame.from_dict(observation,orient ='index')
    temp_df         = temp_df[['Product_Name','Generic_Name','Form', 'API_with_strength' ,'Minimum_Batch_size_NOS',
                       'Minimum_Batch_size_MG','MRDD','LRDD_MG','LRDD_NOS','PDE_VALUE','LRD50','NOEL']]

    store_location = "static/inputData/"+"product_Details.xlsx"
    final_working_directory=MYDIR + "/" +store_location
    temp_df.to_excel(final_working_directory,index=False)
    d = {"error":"none",}
   
    return json.dumps(d)

@app.route("/submit_data")    
def submit_data():
   
    data            = request.args.get('params_data')
    data            = json.loads(data)   
    temp_df         = pd.DataFrame(data)
    

    product_frame  = pd.read_excel(os.path.join(app.config['UPLOAD_FOLDER_INPUTDATA'],"product_Details.xlsx"))   
    file_name = "Cleaning_room_report_{}.xlsx".format(str(datetime.datetime.today().strftime('%d_%m_%Y')))    
    store_location = "static/inputData/"+file_name
    final_working_directory=MYDIR + "/" +store_location
    #final_working_directory=store_location
    
    
   
        
    wb = Workbook()
    ws = wb.active 
    wb = Sheet_Generation.create_equipment_sheet(wb,ws,temp_df)
    wb = Sheet_Generation.create_product_sheet(wb,ws,product_frame)
    wb = Sheet_Generation.create_pde_sheet(wb,ws,product_frame)
    wb = Sheet_Generation.create_toxicity_sheet(wb,ws,product_frame)
    wb = Sheet_Generation.create_dose_base_sheet(wb,ws,product_frame)
    wb.save(final_working_directory)
    
    
    

    if sent_mail:
        send_mail(subject,text,final_working_directory,file_name) 
    d = {"error":"none",
         "file_name":file_name,
         "file_path":store_location}
   
    return json.dumps(d)

if __name__ == '__main__':
    #app.debug = True
    app.run()

